import os
import pandas as pd
from datetime import datetime
from .conect import ConectNvx
from .s_mail import Mensajes, MailConfig

# ==============================================================================
# CORE DE FILTROS Y PROCESAMIENTO DE DATOS
# =============================================================================

class ApiFilter:
    """Esta clase se encarga de manejar la consulta a navixy,
    aplicar filtros a los datos obtenidos, procesar la información y
    exportarla a Excel o enviarla por correo electrónico."""

    ACTIVOS = {"true", "1"}
    INACTIVOS = {"false", "0"}

    def __init__(self, usuario, clave):
        self.con = ConectNvx(usuario, clave)
        self.mail = Mensajes()
        self.mail_config = MailConfig()

    # STATUS USERS ============================================================
    def status_users(self):
        """Consulta la lista de usuarios desde navixy,
        aplica filtros según el estado (activos, inactivos o todos) y muestra los resultados."""

        df = self.con.get_users()

        if self._is_empty(df, "<< No hay usuarios disponibles >>"):
            return

        opcion = input("1- Activos | 2- Inactivos | 3- Todos: ").strip()
        df = self.filtrar_usuarios(df, opcion)

        if self._is_empty(df, "<< No hay registros que cumplan el filtro >>"):
            return

        print(df.to_string(index=True))
        self.export_prompt(df)

    # STATUS PANEL ================================================================
    def status_panel(self):
        """Consulta la lista de trackers desde navixy, muestra los resultados y permite aplicar un filtro de offline."""
        df = self.con.get_trackers()

        if self._is_empty(df, "<< No hay trackers disponibles >>"):
            return

        self.procesar_status(df)

    # STATUS ACCOUNT ==============================================================
    def status_account(self):
        """Consulta la lista de trackers desde navixy, permite filtrar por cuenta,
        muestra los resultados y permite aplicar un filtro de offline."""

        cuenta = input(">> Nombre de la cuenta: ").strip().lower()

        df = self.con.get_trackers()

        if self._is_empty(df, "<< No hay trackers disponibles >>"):
            return

        if "owner_name" not in df.columns:
            print("<< Columna 'owner_name' no encontrada >>")
            return

        df = df[df["owner_name"].str.lower().str.contains(cuenta, na=False)]

        if self._is_empty(df, "<< No hay trackers para esa cuenta >>"):
            return

        self.procesar_status(df)

    # PROCESAR STATUS PANEL O ACCOUNT ==========================================
    def procesar_status(self,df: pd.DataFrame):
        """Procesa el DataFrame de trackers, mostrando los resultados y permitiendo aplicar un filtro de offline."""

        if self._is_empty(df):
            return

        opcion = input(">> Filtro offline (s/n)?: ").lower()

        if opcion == "s":
            df = self.filtrar_trackers_offline(df)

            if self._is_empty(df):
                return
        else:
            if "days_offline" in df.columns:
                df = df.drop(columns=["days_offline"])

        print(df.to_string())
        self.export_prompt(df)

    # EXPORTAR RESULTADOS A EXCEL ======================================================
    def export_prompt(self, df):
        """Muestra un prompt para exportar los resultados a Excel o enviarlos por correo electrónico."""
        if input(">> Exportar (s/n): ").lower() == "s":
            self.export_excel(df)

    def export_excel(self, df):
        """Exporta el DataFrame a un archivo Excel, con opciones para
        guardarlo localmente o enviarlo por correo electrónico."""
        import re
        from .menu import SESSION_LABEL

        # Evitar mutar el original
        df = df.copy()

        def limpiar_valor(valor):
            if isinstance(valor, str):
                return re.sub(r'[\x00-\x1F\x7F-\x9F]', '', valor)
            return valor

        for col in df.select_dtypes(include=["object", "string"]).columns:
            df[col] = df[col].map(limpiar_valor)

        archivo = f"Reporte_{SESSION_LABEL} - {datetime.now():%Y-%m-%d_%H-%M}.xlsx"

        print("<< Guardar archivo >>")
        print("1- Local")
        print("2- Enviar por correo")

        opcion = input(">>: ")

        if opcion == "1":
            ruta = self.mail_config.info.get("RUTA", "")

            if not ruta:
                ruta = input("Ruta (ej: C:/home/user/reportes): ").strip()
                self.mail_config.info["RUTA"] = ruta
                self.mail_config.guardar_config()

            ruta_final = os.path.join(ruta, archivo)

            df.to_excel(ruta_final, index=False)
            self._formatear_excel(ruta_final)

            print(f"<< Archivo guardado: {ruta_final} >>")
            input("ENTER para continuar...")

        elif opcion == "2":

            df.to_excel(archivo, index=False)
            self._formatear_excel(archivo)

            if self.mail_config.mail_config_incompleta():
                self.mail_config.configurar_mail()

            self.mail.crear_mensaje(archivo)
            self.mail.send()

            input("ENTER para continuar...")

    def new_ruta(self):
        """Permite al usuario configurar una nueva ruta para guardar los archivos Excel,
        actualizando la configuración en el archivo JSON."""

        new = input("Nueva ruta (ej: C:/home/user/reportes): ").strip()
        self.mail_config.info["RUTA"] = new
        self.mail_config.guardar_config()

    def cerrar(self):
        """Cierra la conexion con la sesion."""
        self.con.session.close()
        print("\nSesión cerrada.")

    #===========================================================================
    # METODOS ESTATICOS
    #===========================================================================

    # HELPERS ============================================================
    @staticmethod
    def _is_empty(df, msg="<< No hay registros >>"):
        if df.empty:
            print(msg)
            return True
        return False

    # FILTRO USUARIOS ==========================================================
    @staticmethod
    def filtrar_usuarios(df, opcion):
        """Permite al usuario filtrar el DataFrame de usuarios por estado (activos, inactivos o todos)
        y devuelve el DataFrame filtrado."""

        if df.empty:
            return df

        if "activated" not in df.columns:
            print("<< Columna 'activated' no encontrada >>")
            return pd.DataFrame()

        df = df.copy()
        df["activated"] = df["activated"].astype(str).str.lower()

        if opcion == "1":
            return df[df["activated"].isin(ApiFilter.ACTIVOS)]

        if opcion == "2":
            return df[df["activated"].isin(ApiFilter.INACTIVOS)]

        if opcion == "3":
            return df

        print("<< Opción inválida >>")
        return pd.DataFrame()

    # FILTRO TRACKERS ============================================================
    @staticmethod
    def filtrar_trackers_offline(df):
        """Permite al usuario filtrar el DataFrame de trackers por días offline,
        solicitando la cantidad de días o meses como criterio de filtrado, y devuelve el DataFrame filtrado."""

        if df.empty:
            return df

        if "days_offline" not in df.columns:
            print("<< Columna 'days_offline' no encontrada >>")
            return pd.DataFrame()

        print("<< Filtrar dias offline >>")
        print("1- Por dias")
        print("2- Por mes")

        opcion = input(">>: ").strip()

        try:
            if opcion == "1":
                limite = int(input(">> Cantidad de días: "))
            elif opcion == "2":
                meses = int(input(">> Cantidad de meses: "))
                limite = meses * 30
            else:
                return pd.DataFrame()

        except ValueError:
            print("<< Valor inválido >>")
            return pd.DataFrame()

        df = df[df["days_offline"] >= limite]
        return df.sort_values("days_offline", ascending=False)

    # FORMATO DE TABLA A DOCUMENTO EXCEL =====================================================
    @staticmethod
    def _formatear_excel(ruta_archivo):
        from openpyxl import load_workbook
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.utils import get_column_letter

        wb = load_workbook(ruta_archivo)
        ws = wb.active

        filas = ws.max_row
        columnas = ws.max_column

        if filas < 2 or columnas < 1:
            return

        ultima_columna = get_column_letter(columnas)
        rango = f"A1:{ultima_columna}{filas}"

        tabla = Table(
            displayName=f"Tabla_{datetime.now():%H%M%S}",
            ref=rango
        )

        estilo = TableStyleInfo(
            name="TableStyleMedium9",
            showRowStripes=True,
            showColumnStripes=False
        )

        tabla.tableStyleInfo = estilo
        ws.add_table(tabla)

        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter

            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[col_letter].width = max_length + 2

        wb.save(ruta_archivo)