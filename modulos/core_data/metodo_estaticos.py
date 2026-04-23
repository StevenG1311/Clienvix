import pandas as pd
from email.message import EmailMessage
from modulos.mail.s_mail import smtp_send

ACTIVOS = {"true", "1"}
INACTIVOS = {"false", "0"}

def validar_smtp(correo, clave, server, puerto, security):
    msg = EmailMessage()
    msg["From"] = correo
    msg["To"] = correo
    msg["Subject"] = "Clienvix, mensaje de prueba"
    msg.set_content("Conexión exitosa")

    print("Probando conexión...")
    return smtp_send(server, puerto, correo, clave, security, msg)

def render_dataframe(df, titulo="Tabla"):
    from rich.console import Console
    from rich.table import Table

    console = Console()
    table = Table(title=f"[bold]{titulo}[/bold]", show_lines=True)

    for col in df.columns:
        table.add_column(str(col))

    for _, row in df.iterrows():
        fila = []
        for col in df.columns:
            valor = row[col]

            if col == "activated":
                activo = str(valor).lower() in ACTIVOS
                fila.append(f"[green]{valor}[/green]" if activo else f"[red]{valor}[/red]")
            else:
                fila.append(str(valor))

        table.add_row(*fila)

    console.print(table)

def is_empty(df, msg="<< No hay registros >>"):
    if df.empty:
        print(msg)
        return True
    return False

def filtrar_usuarios(df, opcion):
    if df.empty:
        return df

    df = df.copy()
    df["activated"] = df["activated"].astype(str).str.lower()

    if opcion == "1":
        return df[df["activated"].isin(ACTIVOS)]
    elif opcion == "2":
        return df[df["activated"].isin(INACTIVOS)]
    elif opcion == "3":
        return df

    print("<< Opción inválida >>")
    return pd.DataFrame()

def filtrar_trackers_offline(df):
    if df.empty:
        return df

    print("1- Por dias\n2- Por mes")
    opcion = input(">>: ").strip()

    try:
        limite = int(input(">> Cantidad: "))
        if opcion == "2":
            limite *= 30
    except:
        return pd.DataFrame()

    return df[df["days_offline"] >= limite].sort_values("days_offline", ascending=False)

def formatear_excel(ruta_archivo):
    from openpyxl import load_workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter
    import uuid

    wb = load_workbook(ruta_archivo)
    ws = wb.active

    filas = ws.max_row
    columnas = ws.max_column

    if filas < 2:
        return

    rango = f"A1:{get_column_letter(columnas)}{filas}"

    tabla = Table(displayName=f"Tabla_{uuid.uuid4().hex[:8]}", ref=rango)
    estilo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)

    tabla.tableStyleInfo = estilo
    ws.add_table(tabla)

    for col in ws.columns:
        max_len = max(len(str(cell.value)) for cell in col if cell.value)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    wb.save(ruta_archivo)